"""Export service — Excel / CSV / JSON."""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Optional

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.exceptions import ExportError
from app.database.models import Job
from app.database.repositories import ExportRepository, JobRepository
from app.utils.file_manager import FileManager
from app.utils.logger import get_logger

logger = get_logger(__name__)


_COLUMNS = [
    "id",
    "title",
    "company",
    "location",
    "country",
    "remote",
    "sponsorship",
    "salary",
    "job_type",
    "experience_level",
    "source",
    "url",
    "posted_date",
    "scraped_at",
    "status",
    "follow_up_date",
    "notes",
    "description",
]


class ExportService:
    """Materialize the ``jobs`` table to disk in a chosen format."""

    def __init__(
        self,
        repo: Optional[JobRepository] = None,
        record_repo: Optional[ExportRepository] = None,
    ) -> None:
        self._repo = repo or JobRepository()
        self._record_repo = record_repo or ExportRepository()

    # ---------------- Public API ----------------
    def export(
        self,
        fmt: str = "xlsx",
        *,
        filters: Optional[dict[str, Any]] = None,
        prefix: str = "jobs",
        jobs: Optional[Iterable[Job]] = None,
    ) -> Path:
        """Export jobs (optionally pre-filtered) and return the file path."""
        filters = filters or {}
        jobs_list = list(jobs) if jobs is not None else self._fetch_jobs(filters)

        if not jobs_list:
            raise ExportError("No jobs available to export.")

        df = self._to_dataframe(jobs_list)
        fmt = fmt.lower()

        if fmt == "xlsx":
            path = self._export_excel(df, prefix)
        elif fmt == "csv":
            path = self._export_csv(df, prefix)
        elif fmt == "json":
            path = self._export_json(jobs_list, prefix)
        else:
            raise ExportError(f"Unsupported export format: {fmt}")

        self._record_repo.record(
            file_path=str(path), fmt=fmt, rows=len(jobs_list), filters=filters
        )
        logger.info(
            "Exported {n} jobs to {p} ({f})", n=len(jobs_list), p=path.name, f=fmt
        )
        return path

    def export_selected(
        self, jobs: list[Job], fmt: str = "xlsx", *, prefix: str = "selected_jobs"
    ) -> Path:
        return self.export(fmt=fmt, prefix=prefix, jobs=jobs)

    # ---------------- Internals ----------------
    def _fetch_jobs(self, filters: dict[str, Any]) -> list[Job]:
        return self._repo.list_jobs(
            keyword=filters.get("keyword"),
            country=filters.get("country"),
            city=filters.get("city"),
            source=filters.get("source"),
            remote_only=filters.get("remote_only", False),
            employment_type=filters.get("employment_type"),
            experience_level=filters.get("experience_level"),
            status=filters.get("status"),
            statuses=filters.get("statuses"),
            scrape_run_id=filters.get("scrape_run_id"),
            has_follow_up=filters.get("has_follow_up", False),
            limit=filters.get("limit", 50_000),
        )

    def export_follow_ups(self, fmt: str = "xlsx") -> Path:
        """Export jobs with active follow-up statuses (bookmarked/applied/interview/offer)."""
        from app.core.constants import FOLLOW_UP_STATUSES

        return self.export(
            fmt=fmt,
            filters={"statuses": FOLLOW_UP_STATUSES},
            prefix="follow_ups",
        )

    @staticmethod
    def _to_dataframe(jobs: list[Job]) -> pd.DataFrame:
        rows = [
            {
                "id": j.id,
                "title": j.title,
                "company": j.company,
                "location": j.location,
                "country": j.country,
                "remote": j.remote,
                "sponsorship": j.sponsorship,
                "salary": j.salary,
                "job_type": j.job_type,
                "experience_level": j.experience_level,
                "source": j.source,
                "url": j.url,
                "posted_date": j.posted_date.isoformat() if j.posted_date else None,
                "scraped_at": j.scraped_at.isoformat() if j.scraped_at else None,
                "status": j.status,
                "follow_up_date": (
                    j.follow_up_date.isoformat() if j.follow_up_date else None
                ),
                "notes": j.notes,
                "description": (j.description or "")[:5_000],
            }
            for j in jobs
        ]
        return pd.DataFrame(rows, columns=_COLUMNS)

    @staticmethod
    def _autosize_columns(worksheet, df: pd.DataFrame) -> None:
        for idx, col in enumerate(df.columns, start=1):
            try:
                series = df[col].astype(str)
                max_len = max(series.map(len).max(), len(str(col))) + 2
                max_len = min(max_len, 60)
            except (ValueError, TypeError):
                max_len = 18
            worksheet.column_dimensions[get_column_letter(idx)].width = max_len

    def _export_excel(self, df: pd.DataFrame, prefix: str) -> Path:
        path = FileManager.export_path(prefix, "xlsx")

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="All Jobs", index=False)

            # Per-source sheets
            for source, sub in df.groupby("source", dropna=True):
                safe = (source or "Unknown")[:28]
                sub.to_excel(writer, sheet_name=safe, index=False)

            wb = writer.book
            header_font = Font(bold=True, color="FFFFFFFF", size=11)
            header_fill = PatternFill("solid", fgColor="FF1F6FEB")
            header_align = Alignment(horizontal="center", vertical="center")

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions
                self._autosize_columns(ws, df)
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align

            meta = wb.create_sheet("Metadata", 0)
            meta["A1"] = "JobHunter Pro Export"
            meta["A1"].font = Font(bold=True, size=14)
            meta["A2"] = f"Generated: {datetime.now().isoformat(timespec='seconds')}"
            meta["A3"] = f"Total rows: {len(df)}"
            meta.column_dimensions["A"].width = 60

        return path

    @staticmethod
    def _export_csv(df: pd.DataFrame, prefix: str) -> Path:
        path = FileManager.export_path(prefix, "csv")
        df.to_csv(path, index=False, encoding="utf-8-sig")
        return path

    @staticmethod
    def _export_json(jobs: list[Job], prefix: str) -> Path:
        path = FileManager.export_path(prefix, "json")
        payload = [j.to_dict() for j in jobs]
        with path.open("w", encoding="utf-8") as fh:
            json.dump(payload, fh, ensure_ascii=False, indent=2, default=str)
        return path
